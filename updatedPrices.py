import openpyxl, os

os.chdir('/home/tabish/Documents/LEARN_PYTHON/additional material for Automate the boring stuff/')

wb = openpyxl.load_workbook('produceSales.xlsx')

sheet = wb['Sheet']

print('Updating sales list...')

#print(sheet['A2'].value)
#print(type(sheet['B2'].value))


# Update Price below

revisedPrice = {'Garlic': 3.07, 'Celery': 1.19, 'Lemon': 1.27}

for i in range(2, sheet.max_row+1): # << Attend to range
    if sheet.cell(row=i, column=1).value in revisedPrice:
        sheet.cell(row=i, column=2).value = revisedPrice[sheet.cell(row=i, column=1).value]
        #print('The price for ' + sheet.cell(row=i, column=1).value + ' is now R' + str(revisedPrice[sheet.cell(row=i, column=1).value]))

print('Price list updated!')

os.chdir('/home/tabish/Documents/LEARN_PYTHON/PyProjects/')

wb.save('updatedProduceSales.xlsx')

